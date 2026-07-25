#!/usr/bin/env python3
"""Inspect the final Test workbook and image directory without modifying them."""

from __future__ import annotations

import argparse
from collections import Counter
import hashlib
import json
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}


class SourceAuditError(ValueError):
    """Raised when the final Test sources cannot be inspected safely."""


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalized_basename(value: Any) -> str:
    text = str(value).strip().replace("\\", "/")
    return PurePosixPath(text).name.casefold()


def inspect_images(image_root: Path) -> tuple[dict, set[str], set[str]]:
    if not image_root.is_dir():
        raise SourceAuditError(f"image directory does not exist: {image_root}")
    images = sorted(
        (
            path
            for path in image_root.rglob("*")
            if path.is_file() and path.suffix.casefold() in IMAGE_SUFFIXES
        ),
        key=lambda path: path.relative_to(image_root).as_posix().casefold(),
    )
    if not images:
        raise SourceAuditError(f"no supported images under: {image_root}")

    names = Counter(path.name.casefold() for path in images)
    stems = Counter(path.stem.casefold() for path in images)
    extensions = Counter(path.suffix.casefold() for path in images)
    unreadable: list[str] = []
    manifest_digest = hashlib.sha256()
    for path in images:
        relative = path.relative_to(image_root).as_posix()
        try:
            with Image.open(path) as image:
                image.verify()
        except (OSError, ValueError) as exc:
            unreadable.append(f"{relative}: {exc}")
        digest = file_sha256(path)
        manifest_digest.update(relative.encode("utf-8"))
        manifest_digest.update(b"\0")
        manifest_digest.update(str(path.stat().st_size).encode("ascii"))
        manifest_digest.update(b"\0")
        manifest_digest.update(digest.encode("ascii"))
        manifest_digest.update(b"\n")

    return (
        {
            "root": str(image_root.resolve()),
            "count": len(images),
            "extension_counts": dict(sorted(extensions.items())),
            "duplicate_basenames": sorted(name for name, count in names.items() if count > 1),
            "duplicate_stems": sorted(name for name, count in stems.items() if count > 1),
            "unreadable_count": len(unreadable),
            "unreadable_examples": unreadable[:10],
            "manifest_sha256": manifest_digest.hexdigest(),
        },
        set(names),
        set(stems),
    )


def compact_value(value: Any, limit: int = 120) -> Any:
    if value is None or isinstance(value, (bool, int, float)):
        return value
    text = str(value).strip()
    return text if len(text) <= limit else text[: limit - 3] + "..."


def inspect_workbook(
    workbook_path: Path,
    image_names: set[str],
    image_stems: set[str],
) -> dict:
    if not workbook_path.is_file():
        raise SourceAuditError(f"workbook does not exist: {workbook_path}")
    try:
        workbook = load_workbook(
            workbook_path,
            read_only=False,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise SourceAuditError(f"cannot open workbook: {exc}") from exc

    sheets: list[dict] = []
    for worksheet in workbook.worksheets:
        preview: list[dict] = []
        formula_count = 0
        nonempty_count = 0
        column_values: dict[int, list[Any]] = {
            column: [] for column in range(1, worksheet.max_column + 1)
        }
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None or (isinstance(value, str) and not value.strip()):
                    continue
                nonempty_count += 1
                column_values[cell.column].append(value)
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                if len(preview) < 30 and cell.row <= 15 and cell.column <= 20:
                    preview.append(
                        {
                            "cell": cell.coordinate,
                            "value": compact_value(value),
                        }
                    )

        profiles: list[dict] = []
        for column, values in column_values.items():
            if not values:
                continue
            strings = [str(value).strip() for value in values]
            normalized_names = [normalized_basename(value) for value in values]
            exact_name_matches = sum(name in image_names for name in normalized_names)
            exact_stem_matches = sum(
                PurePosixPath(name).stem.casefold() in image_stems
                for name in normalized_names
            )
            counts = Counter(strings)
            profiles.append(
                {
                    "column": get_column_letter(column),
                    "nonempty": len(values),
                    "unique": len(counts),
                    "top_values": [
                        {"value": compact_value(value, 80), "count": count}
                        for value, count in counts.most_common(8)
                    ],
                    "image_basename_matches": exact_name_matches,
                    "image_stem_matches": exact_stem_matches,
                }
            )

        sheets.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "nonempty_cells": nonempty_count,
                "formula_cells": formula_count,
                "merged_ranges": [str(value) for value in worksheet.merged_cells.ranges],
                "preview": preview,
                "column_profiles": profiles,
            }
        )
    workbook.close()
    return {
        "path": str(workbook_path.resolve()),
        "size_bytes": workbook_path.stat().st_size,
        "sha256": file_sha256(workbook_path),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def audit_sources(workbook_path: Path, image_root: Path) -> dict:
    image_summary, image_names, image_stems = inspect_images(image_root)
    workbook_summary = inspect_workbook(workbook_path, image_names, image_stems)
    return {
        "protocol_version": "final_test250_source_inspection_v1",
        "source_only": True,
        "model_inference_run": False,
        "workbook": workbook_summary,
        "images": image_summary,
        "status": "PASS",
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--image-root", required=True, type=Path)
    args = parser.parse_args()
    try:
        summary = audit_sources(args.workbook, args.image_root)
    except (SourceAuditError, OSError) as exc:
        print(f"ERROR: {exc}")
        return 2
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    print("FINAL_TEST250_SOURCE_INSPECTION: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
